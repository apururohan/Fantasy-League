import tkinter as r
import tkinter.font as tkfont
question2=r.Tk()
question2.title("Question 2")
question2.configure(bg='cyan')
question2.geometry('770x500')

labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=40)

l=r.Label(question2,font=labelsty,text='Question  2')
l.pack(side='top')
l.place()
l.configure(bg='cyan')

f1=tkfont.Font(family="Lucida Grande",size=10)
f1.configure(size=20)

f2=tkfont.Font(family="Lucida Grande",size=10)
f2.configure(size=15)

f3=tkfont.Font(family='Lucida Grande',size=10)
f3.configure(size=10)

l1=r.Label(question2,font=f1,text='Who will take more number of catches')
l1.place(x=140,y=170)
l1.configure(bg='cyan')

e1=r.Entry(question2,font=f1,width=11)
e1.place(x=167,y=220)

e2=r.Entry(question2,font=f1,width=11)
e2.place(x=405,y=220)



def nextq():
    player1=e1.get()
    player2=e2.get()
    if(player1==player2):
        l2=r.Label(question2,font=f3,text='Same player name cannot be given twice.')
        l2.place(x=250,y=140)
        l2.configure(bg='cyan')
    else:
        import openpyxl
        wb=openpyxl.load_workbook("Project.xlsx")
        sh1=wb['Sheet1']
        c=sh1.max_column
        name=sh1.cell(1,c-1).value
        match=""
        for i in name:
            if(i=='%' or i=='1' or i=='2'):
                pass
            else:
                match=match+i
        match=match[:-1]
        match=match+'.txt'
        f = open(match,'a+')
        writeinto=['%','\n',player1,'\n','%','\n',player2,'\n']
        f.writelines(writeinto)
        f.close()
        question2.destroy()
        import question3


b1=r.Button(question2,text='Next',font=f2,command=nextq)
b1.place(x=343,y=270)
