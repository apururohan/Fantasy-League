import tkinter as r
import tkinter.font as tkfont
existing=r.Tk()
existing.title("EXISTING USER")
existing.configure(bg='cyan')
existing.geometry('770x500')

labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=100)

l=r.Label(existing,text='Login User',font=labelsty)
l.place(x=60)
l.configure(bg='cyan')

f1=tkfont.Font(family="Lucida Grande",size=10)
f1.configure(size=25)

l1=r.Label(existing,text='Enter the username',font=f1)
e1=r.Entry(existing,width=25,font=f1)
l2=r.Label(existing,text='Enter the password',font=f1)
e2=r.Entry(existing,width=25,font=f1)

l1.configure(bg='cyan')
l2.configure(bg='cyan')

l1.place(y=250)
e1.place(x=300,y=250)
l2.place(y=300)
e2.place(x=300,y=300)

f2=tkfont.Font(family='Lucida Grande',size=10)
f2.configure(size=20)

f3=tkfont.Font(family='Lucida Grande',size=10)
f3.configure(size=10)

def back():
    existing.destroy()
    import user
def login():
    import openpyxl
    wb=openpyxl.load_workbook("Project.xlsx")
    sh1=wb['Sheet1']
    user=[]
    row=sh1.max_row
    username=e1.get()
    password=e2.get()
    for i in range(2,row+1):
        if(sh1.cell(i,1).value==username):
            pos=i
    print(pos)
    print(sh1.cell(pos,2).value)
    print(password)
    if(int(sh1.cell(pos,2).value)==int(password)):
        print("Success is not a destination.It's a journey")
        f = open("user.txt",'w')
        f.write(username)
        f.write('\n')
        f.write(str(pos))
        f.close()
        existing.destroy()
        import home
    else:
        l3=r.Label(existing,text='Entered username and password are incorrect.')
        l3.place(x=300,y=205)
        
b1=r.Button(existing,text='Login',font=f2,command=login)
b1.place(x=350,y=357)
b2=r.Button(existing,text='Back',font=f3,command=back)
b2.place(x=15,y=450)
existing.mainloop()
