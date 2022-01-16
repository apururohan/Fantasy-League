import tkinter as r
import tkinter.font as tkfont
import re
import openpyxl
bet=r.Tk()
bet.title("PLAY")
bet.configure(bg='cyan')


labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=100)


f1=tkfont.Font(family="Lucida Grande",size=10)
f1.configure(size=25)

def match():
    filename=name+'.txt'
    f1=open('name.txt','w')
    f1.write(filename)
    f1.close()
    print(filename)
    import sorry
            

wb=openpyxl.load_workbook("Project.xlsx")
sh1=wb['Sheet1']
col=sh1.max_column
ro=sh1.max_row
matches=[]
l=0
s=r.IntVar()
s.set("-1")
for i in range(4,col,8):
    newstr=""
    present=sh1.cell(1,i).value
    for k in range(0,len(present)-1):
        newstr=newstr+present[k]
    for j in range(2,ro+1):
        if(sh1.cell(j,i).value=='$'):
            break
    if(j==ro):
        name_list=[]
        #print(present)
        #print(newstr)
        name_list=re.findall('[A-Z][^A-Z]*',newstr)
        print(name_list)
        name1=name_list[0]
        name1=name1[:-1]
        name2=name_list[1]
        #print(name1)
        #print(name2)
        label_text=name1+" v/s "+name2
        matches.append(name1+name2)
        #b=r.Button(bet,text=label_text,font=f1,command=match(name1+name2,i))
        la=r.Label(bet,text=label_text,font=f1)
        la.pack()
label=r.Label(bet,text="Enter the name of the available matches shown above :",font=f1)
label.pack()
entry=r.Entry(bet,width=15,font=f1)
entry.pack()
name=entry.get()
bo=r.Button(bet,text="Next",font=f1,command=match())
bo.pack()

