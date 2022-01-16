import tkinter as r
import tkinter.font as tkfont
prematch=r.Tk()
prematch.title("Prematch setup")
prematch.configure(bg='cyan')
prematch.geometry('770x500')

labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=40)

l=r.Label(prematch,font=labelsty,text='Select two teams')
l.pack(side='top')
l.place()
l.configure(bg='cyan')

f1=tkfont.Font(family="Lucida Grande",size=10)
f1.configure(size=15)

f2=tkfont.Font(family="Lucida Grande",size=10)
f2.configure(size=15)

f3=tkfont.Font(family='Lucida Grande',size=10)
f3.configure(size=10)

l1=r.Label(prematch,font=f1,text='First team')
l1.place(x=250,y=200)
l1.configure(bg='cyan')

e1=r.Entry(prematch,width=11,font=f1)
e1.place(x=377,y=200)

l2=r.Label(prematch,font=f1,text='Second team')
l2.place(x=250,y=250)
l2.configure(bg='cyan')

e2=r.Entry(prematch,width=11,font=f1)
e2.place(x=377,y=250)

def create():
    first=e1.get()
    second=e2.get()
    if(first==second):
        l3=r.Label(prematch,font=f1,text='The match cannot be setup between two same teams.')
        l3.place(x=105,y=100)
        l3.configure(bg='cyan')
    else:
        import openpyxl
        wb=openpyxl.load_workbook("Project.xlsx")
        sh1=wb['Sheet1']
        c=sh1.max_column
        name1=first+'%'+second+'1'
        name2=first+'%'+second+'2'
        name3=first+'%'+second+'3'
        name4=first+'%'+second+'4'
        name5=first+'%'+second+'5'
        name6=first+'%'+second+'6'
        name7=first+'%'+second+'7'
        name8=first+'%'+second+'8'
        sh1.cell(row=1,column=c+1,value=name1)
        sh1.cell(row=1,column=c+2,value=name2)
        sh1.cell(row=1,column=c+3,value=name3)
        sh1.cell(row=1,column=c+4,value=name4)
        sh1.cell(row=1,column=c+5,value=name5)
        sh1.cell(row=1,column=c+6,value=name6)
        sh1.cell(row=1,column=c+7,value=name7)
        sh1.cell(row=1,column=c+8,value=name8)
        sh1.cell(row=2,column=c+7,value="#")
        sh1.cell(row=2,column=c+6,value="#")
        sh1.cell(row=2,column=c+5,value="#")
        sh1.cell(row=2,column=c+4,value="#")
        sh1.cell(row=2,column=c+3,value="#")
        sh1.cell(row=2,column=c+2,value="#")
        sh1.cell(row=2,column=c+1,value="#")
        sh1.cell(row=2,column=c+8,value="#")
        wb.save('Project.xlsx')
        prematch.destroy()
        import question1
def back():
    prematch.destroy()
    import match
b2=r.Button(prematch,text='Back',font=f3,command=back)
b2.place(x=15,y=450)

b1=r.Button(prematch,text='Set',font=f2,command=create)
b1.place(x=350,y=300)
