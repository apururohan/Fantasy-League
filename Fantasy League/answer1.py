import tkinter as r
import tkinter.font as tkfont
answer1=r.Tk()
answer1.title("ANSWER THE FOLLOWING")
answer1.configure(bg='cyan')
answer1.geometry('770x500')

labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=100)

l=r.Label(answer1,text='Question1',font=labelsty)
l.place(x=60)
l.configure(bg='cyan')

f1=tkfont.Font(family="Lucida Grande",size=10)
f1.configure(size=20)


fp1=open("name.txt","r")
filename=fp1.read()
fp1.close()
#print(filename)
fp1=open(filename,"r")
player=fp1.read()
players=player.split("\n")
#print(players)
fp1.close()
fp1=open("user.txt","r")
fp1.readline()
pos=int(fp1.readline())
fp1.close()
l1=r.Label(answer1,text="Who will hit the more number of sixes",font=f1)
l1.configure(bg='cyan')
l2=r.Label(answer1,text=players[1],font=f1)
e2=r.Entry(answer1,width=5,font=f1)
l3=r.Label(answer1,text=players[3],font=f1)
e3=r.Entry(answer1,width=5,font=f1)
l2.configure(bg='cyan')
l3.configure(bg='cyan')

l1.place(x=155,y=190)
l2.place(x=215,y=225)
l3.place(x=410,y=225)
e2.place(x=215,y=260)
e3.place(x=430,y=260)

def nextquestion():
    option1=int(e2.get())
    option2=int(e3.get())
    #print(option1+option2)
    if(option1+option2==100):
        import openpyxl
        wb=openpyxl.load_workbook("Project.xlsx")
        sh1=wb['Sheet1']
        matchname=filename[0]
        #print(filename)
        for i in range(1,len(filename)):
            if(ord(filename[i])>=97):
                matchname=matchname+filename[i]
            elif(ord(filename[i])==46):
                continue
            else:
                matchname=matchname+'%'+filename[i]
        matchname=matchname[:-3]
        matchname=matchname+'1'
        print(matchname)
        col=sh1.max_column
        for i in range(4,col+1,1):
            if(matchname==sh1.cell(1,i).value):
                break
            else: pass
        fp=open("answer.txt","w")
        fp.write(str(i))
        fp.close()
        print(i)
        sh1.cell(row=pos,column=i,value=option1)
        sh1.cell(row=pos,column=i+1,value=option2)
        wb.save("Project.xlsx")
        answer1.destroy()
        import answer2
    else:
        l4=r.Label(answer1,text="Enter a sum of 100",font=f1)
        l4.configure(bg='red')
        l4.place(x=260,y=150)

b1=r.Button(answer1,text="NEXT",font=f1,command=nextquestion)
b1.place(x=315,y=310)
