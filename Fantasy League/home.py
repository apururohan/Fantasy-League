import openpyxl
import tkinter as r
import tkinter.font as tkfont


def back():
    home.destroy()
    import user

def leaderboard():
    import leaderboard

def bet():
    home.destroy()
    import bet1

def logoutc():
    home.destroy()
    import user

f =open("user.txt","r")
username = f.readline()
index= int(f.readline())
f.close()
print(username)


home=r.Tk()
home.title("HOME")
home.configure(bg='cyan')
home.geometry('770x500')

labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=100)



l=r.Label(home,text='HOME',font=labelsty)
l.place(x=185)
l.configure(bg='cyan')

f1=tkfont.Font(family="Lucida Grande",size=10)
f1.configure(size=25)

f4=tkfont.Font(family="Lucida Grande",size=10)
f4.configure(size=15)

f3=tkfont.Font(family="Lucida Grande",size=10)
f3.configure(size=10)

b1=r.Button(home,text='Leaderboard',font=f1,command=leaderboard)
b1.pack()
b1.place(x=275,y=220)

b2=r.Button(home,text='Bet',font=f1,command=bet)
b2.pack()
b2.place(x=342,y=300)

b2=r.Button(home,text='Back',font=f3,command=back)
b2.place(x=15,y=450)



wb = openpyxl.load_workbook("Project.xlsx")
sh1= wb['Sheet1']
userid=[]
rows=sh1.max_row
'''
ans=sh1.cell(index,3).value
l1=r.Label(home,text=ans,font=f4)
l1.place(x=500,y=100)
b3=r.Button(home,text="Logout",font=f4,command=logoutc)
b3.place(x=500,y=120)'''

frame=r.Frame(home)
frame.configure(bg='yellow')
login=r.Label(frame,text=str(sh1.cell(index,1).value),font=f4)
login.configure(bg='yellow')
points=r.Label(frame,text='Points : '+str(sh1.cell(index,3).value),font=f4)
points.configure(bg='yellow')
login.grid(row=1,column=1)
points.grid(row=2,column=1)
logout=r.Button(frame,text='logout',font=f4,command=logoutc)
logout.grid(row=3,column=1)
logout.configure(bg='violet')
frame.place(x=650,y=25)
#home.mainloop()


