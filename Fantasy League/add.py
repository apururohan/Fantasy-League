import tkinter as r
import tkinter.font as tkfont
bonus=r.Tk()
bonus.title("Bonus")
bonus.configure(bg='cyan')
bonus.geometry('770x500')

labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=50)

l=r.Label(bonus,font=labelsty,text='MATCH')
l.pack(side='top')
l.place()
l.configure(bg='cyan')

f1=tkfont.Font(family="Lucida Grande",size=10)
f1.configure(size=25)

def enter():
    add=e1.get()
    import openpyxl
    wb=openpyxl.load_workbook("Project.xlsx")
    sh1=wb['Sheet1']
    row=sh1.max_row
    col=sh1.max_column
    for i in range(3,row+1):
        ans=sh1.cell(i,3).value
        ans=ans+int(add)
        sh1.cell(row=i,column=3,value=ans)
    wb.save("Project.xlsx")
    bonus.destroy()
    print("Program terminated successfully.")

l1=r.Label(bonus,font=f1,text='Enter the bonus amount')
l1.place(x=100,y=200)
l1.configure(bg='cyan')

e1=r.Entry(bonus,font=f1,width=10)
e1.place(x=455,y=200)

b1=r.Button(bonus,font=f1,text="Enter",command=enter)
b1.place(x=320,y=300)

f3=tkfont.Font(family='Lucida Grande',size=10)
f3.configure(size=10)

def back():
    bonus.destroy()
    import matchsetup

b2=r.Button(bonus,text='Back',font=f3,command=back)
b2.place(x=15,y=450)
