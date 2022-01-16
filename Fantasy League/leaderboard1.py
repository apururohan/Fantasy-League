#leaderboard.configure(bg='cyan')
#leaderboard.geometry('770x500')


import openpyxl
wb=openpyxl.load_workbook("Project.xlsx")
sh1=wb['Sheet1']
row=sh1.max_row
col=sh1.max_column
inf=[]
ro=0
c=0
pos=-1
for i in range(3,row+1):
    temp=[]
    ro=ro+1
    for j in range(1,col+1):
        temp.append(sh1.cell(i,j).value)
    inf.append(temp)
print(inf)
print("\n\n\n\n")
'''for i in range(0,ro+1):
    for j in range(i,ro+1):
        if(int(inf[i][2])<int(inf[j][2])):
            pos=i
    if(pos!=i):
        temp=inf[i][:]
        inf[j]=inf[i]
        inf[i]=temp
'''

sorted_inf=sorted(inf ,key=lambda x : x[2])
print(sorted_inf)
sorted_inf.reverse()

for i in range(3,row+1):
    for j in range(1,col+1):
        sh1.cell(row=i,column=j,value=sorted_inf[i-3][j-1])
wb.save("Project.xlsx")

import leaderboard
