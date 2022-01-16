import tkinter as r
import tkinter.font as tkfont
new=r.Tk()
new.title("NEW USER")
new.configure(bg='cyan')
new.geometry('770x500')

labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=100)

l=r.Label(new,text='New User',font=labelsty)
l.place(x=85)
l.configure(bg='cyan')

f1=tkfont.Font(family="Lucida Grande",size=10)
f1.configure(size=25)

l1=r.Label(new,text='Enter the username',font=f1)
e1=r.Entry(new,width=25,font=f1)
l2=r.Label(new,text="Enter the password",font=f1)
e2=r.Entry(new,width=25,font=f1)

l1.configure(bg='cyan')
l2.configure(bg='cyan')

l1.place(y=250)
e1.place(x=300,y=250)
l2.place(y=300)
e2.place(x=300,y=300)


f2=tkfont.Font(family="Lucida Grande",size=10)
f2.configure(size=20)

f3=tkfont.Font(family="Lucida Grande",size=10)
f3.configure(size=10)

def back():
    new.destroy()
    import user
def create():
    import openpyxl
    wb=openpyxl.load_workbook("Project.xlsx")
    sh1=wb['Sheet1']
    user_ids=[]
    row=sh1.max_row
    username=e1.get()
    password=e2.get()
    for i in range(2,row+1):
        user_ids.append(sh1.cell(i,1).value)
    #print(username)
    #print(user_ids)
    if username in user_ids:
        l3=r.Label(new,text='Sorry! The username already exists',font=f3)
        l3.place(x=300,y=205)
    else:
        sh1.cell(row=row+1,column=1,value=username)
        sh1.cell(row=row+1,column=2,value=password)
        sh1.cell(row=row+1,column=3,value=500)
        wb.save('Project.xlsx')
        new.destroy()
        import existingdemo

b1=r.Button(new,text='Create',font=f2,command=create)
b1.place(x=350,y=357)
b2=r.Button(new,text='Back',font=f3,command=back)
b2.place(x=15,y=450)
new.mainloop()

