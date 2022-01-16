import tkinter as r
import tkinter.font as tkfont
leaderboard=r.Tk()
leaderboard.title("LEADERBOARD")
#leaderboard.configure(bg='cyan')
#leaderboard.geometry('770x500')

labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=100)

import openpyxl
wb=openpyxl.load_workbook("Project.xlsx")
sh1=wb['Sheet1']
row=sh1.max_row
r.Label(leaderboard,text="Name of the user").grid(row=0,column=0)
r.Label(leaderboard,text="Points").grid(row=0,column=1)
for i in range(3,row+1):
    r.Label(leaderboard,text=sh1.cell(i,1).value).grid(row=i-1,column=0)
    r.Label(leaderboard,text=sh1.cell(i,3).value).grid(row=i-1,column=1)

leaderboard.mainloop()
