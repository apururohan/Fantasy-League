import tkinter as r
import tkinter.font as tkfont
import openpyxl

start=r.Tk()


f2=open('user.txt','r')
username=f2.readline()
pos=int(f2.readline())
f2.close()

fo1=tkfont.Font(family="Lucida Grande",size=10)
fo1.configure(size=25)

wb1=openpyxl.load_workbook("Project.xlsx")
sh1=wb1['Sheet1']
row=sh1.max_row
val=sh1.cell(pos,3).value
if(val<400):
    l=r.Label(start,text="Sorry! No enough funds available.",font=fo1)
    l.pack()
    start.destroy()
    exit(0)
else:
    start.destroy()
    import answer1
