import openpyxl
wb=openpyxl.load_workbook("Project.xlsx")
sh1=wb['Sheet1']
fp1=open("name.txt","r")
filename=fp1.read()
fp1.close()
matchname=filename[0]
for i in range(1,len(filename)):
    if(ord(filename[i])>=97):
        matchname=matchname+filename[i]
    elif(ord(filename[i])==46):
        continue
    else:
        matchname=matchname+'%'+filename[i]
matchname=matchname[:-3]
matchname=matchname+'1'
print(matchname)
col=sh1.max_column
row=sh1.max_row
for i in range(2,row):
    for j in range(4,col):
        if(sh1.cell(i,j).value==None ):
            sh1.cell(row=i,column=j,value=0)
wb.save("Project.xlsx")

for i in range(1,col):
    if(sh1.cell(1,i).value==matchname):
        pos=i
for i in range(3,row+1):
    ans=sh1.cell(i,3).value
    print(ans)
    for j in range(pos,pos+8):
        print(sh1.cell(i,j).value)
        if(not (sh1.cell(i,j).value==0)):
            ans=ans+(2*(int(sh1.cell(i,j).value)))
        else:
            ans=ans-(sh1.cell(i,j).value)/2
sh1.cell(row=i,column=3,value=ans)
wb.save("Project.xlsx")
