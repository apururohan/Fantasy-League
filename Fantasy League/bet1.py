import tkinter as r
import tkinter.font as tkfont
import re
import openpyxl
bet=r.Tk()
bet.title("PLAY")
bet.configure(bg='cyan')

labelsty=tkfont.Font(family="Lucida Grande",size=10)
labelsty.configure(size=100)

f1=tkfont.Font(family="Lucida Grande",size=10)
f1.configure(size=25)

def rohan():
    roshan=ent.get()
    filename=roshan+'.txt'
    f1=open('name.txt','w')
    f1.write(filename)
    f1.close()
    #print(filename)
    bet.destroy()
    import sorry
    

wb=openpyxl.load_workbook("Project.xlsx")
sh1=wb['Sheet1']
col=sh1.max_column
ro=sh1.max_row
matches=[]
l=0

for i in range(4,col,8):
    if(sh1.cell(2,i).value=='#'):
        newstr=""
        present=sh1.cell(1,i).value
        for k in range(0,len(present)-1):
            newstr=newstr+present[k]
        name_list=[]
        name_list=re.findall('[A-Z][^A-Z]*',newstr)
        name1=name_list[0]
        name1=name1[:-1]
        name2=name_list[1]
        matches.append(name1+name2)
        la=r.Label(bet,text=name1+' v/s '+name2,font=f1)
        la.pack()
label=r.Label(bet,text="Enter the names of the two teams",font=f1)
label.pack()
ent=r.Entry(bet,font=f1)
ent.pack()
bu=r.Button(bet,text="Proceed",command=rohan)
bu.pack()

